import os
import sys
sys.path.append(".")
from xactron_parser import XactronParser
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class ExportExcel():
    def __init__(self, path, data):
        self.path = path
        parser = XactronParser(data, None, False)
        self.component = parser.component
        self.memoryMaps = self.component.memoryMaps.memoryMap
        self.projectName = self.component.library
        self.finalPath = os.path.join(self.path, self.projectName + "." + "xlsx")
        self.book = Workbook()
        self.sheets = self.book.sheetnames
        self.regCells = []
        self.regCellsForAppend = []
        self.cleanSheets()
        self.styles()


    def write(self):
        self.createFuncSheet()
        self.createMappingSheet()
        self.book.save(self.finalPath)


    def combine(self):
        pass
        # TODO COMBINE Excel


    def styles(self):
        self.mediumBorder = Border(left=Side(style='medium'), 
                    right=Side(style='medium'), 
                    top=Side(style='medium'), 
                    bottom=Side(style='medium'))

        self.thinBorder = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))


    def cleanSheets(self):
        for sheet in self.sheets:
            if sheet == "Sheet":
                self.book.remove(self.book['Sheet'])
            if sheet == "Feuil1":
                self.book.remove(self.book['Feuil1'])


    def setBorder(self, ws, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
        rows = ws[cell_range]

        for row in rows:
            for cell in row:
                cell.border = border    


    def createFuncSheet(self):
        for memoryMap in self.memoryMaps:
            for addressSpace in memoryMap.addressBlock:
                if addressSpace.name in self.book.sheetnames:
                    funcSheet = self.book[addressSpace.name]
                    loop = 0
                    while funcSheet.merged_cells:
                        loop += 1
                        for cellrange in funcSheet.merged_cells:
                            if cellrange.max_col < 8:
                                funcSheet.merged_cells.remove(str(cellrange))
                                loop = 0
                        if loop == 5:
                            break
                else:
                    funcSheet = self.book.create_sheet(addressSpace.name)
                
                funcSheet.sheet_view.zoomScale = 85
                registerVStart = 4
                funcSheet.column_dimensions['A'].width = 11
                funcSheet.column_dimensions['B'].width = 10
                funcSheet.column_dimensions['C'].width = 5
                funcSheet.column_dimensions['D'].width = 13
                funcSheet.column_dimensions['E'].width = 10
                funcSheet.column_dimensions['F'].width = 20
                funcSheet.column_dimensions['G'].width = 100
                funcSheet.column_dimensions['X'].width = 158

                if len(addressSpace.register) > 0:
                    for register in sorted(addressSpace.register, key=lambda register: register.addressOffset):
                        fieldVStart = registerVStart + 3
                        funcSheet.merge_cells(start_row=registerVStart, start_column=2, end_row= registerVStart, end_column=6)
                        #funcSheet.row_dimensions[registerVStart] = 5
                        funcSheet.cell(row=registerVStart, column=2).value = register.name
                        funcSheet.cell(row=registerVStart, column=7).value ="0x{} + 0x{}".format(hex(addressSpace.baseAddress)[2:].upper(), hex(register.addressOffset)[2:].upper())
                        funcSheet.cell(row=registerVStart, column=7).font = Font(size=14)
                        funcSheet.cell(row=registerVStart, column=7).alignment = Alignment(horizontal='right')
                        funcSheet.cell(row=registerVStart, column=7).fill = PatternFill("solid", fgColor="EDEDED")
                        funcSheet.cell(row=registerVStart, column=2).fill = PatternFill("solid", fgColor="8EA9DB")
                        funcSheet.cell(row=registerVStart, column=2).font = Font(bold=True, size=14)
                        funcSheet.cell(row=registerVStart+1, column=2).value = register.description
                        funcSheet.cell(row=registerVStart+1, column=24).font = Font(bold=True, size=14, color="FFFFFF")
                        funcSheet.cell(row=registerVStart+1, column=24).value = register.description
                        funcSheet.merge_cells(start_row=registerVStart + 1, start_column=2, end_row=registerVStart + 1, end_column=7)
                        funcSheet.cell(row=registerVStart+1, column=2).alignment = Alignment(wrapText=True)
                        funcSheet.cell(row=registerVStart+1, column=24).alignment = Alignment(wrapText=True)
                        funcSheet.cell(row=registerVStart+1, column=2).font = Font(bold=True, size=14)
                        funcSheet.cell(row=registerVStart+1, column=2).fill = PatternFill("solid", fgColor="EDEDED")
                        funcSheet.cell(row=registerVStart+2, column=2).value = "Bits"
                        funcSheet.cell(row=registerVStart+2, column=3).value = "Size"
                        funcSheet.cell(row=registerVStart+2, column=4).value = "Default Value"
                        funcSheet.cell(row=registerVStart+2, column=5).value = "Access"
                        funcSheet.cell(row=registerVStart+2, column=6).value = "Name"
                        funcSheet.cell(row=registerVStart+2, column=7).value = "Description"
                        for row in range(2,8):
                            funcSheet.cell(row=registerVStart+2, column=row).fill = PatternFill("solid", fgColor="EDEDED")
                            funcSheet.cell(row=registerVStart+2, column=row).font = Font(bold=True, size=11)
                            funcSheet.cell(row=registerVStart+2, column=row).alignment = Alignment(wrapText=True)

                        for field in reversed(sorted(register.field, key=lambda field: field.value)):
                            posh = field.bitOffset
                            posl = posh + field.bitWidth - 1
                            fieldBit = "[{}:{}]".format(posh, posl)
                            size = int(posh) - int(posl) + 1
                            # fieldAccess = accessFormatorBoolToString(field.accessRead, field.accessWrite, field.accessWriteOnce, field.accessReadOnce)
                            funcSheet.cell(row=fieldVStart, column=2).value = fieldBit
                            funcSheet.cell(row=fieldVStart, column=3).value = int(size)
                            funcSheet.cell(row=fieldVStart, column=4).value = hex(field.value).upper()
                            # funcSheet.cell(row=fieldVStart, column=5).value = fieldAccess
                            funcSheet.cell(row=fieldVStart, column=6).value = field.name
                            if len(field.enumeratedValues) > 0:
                                evs_string = ""
                                for evs in field.enumeratedValues:
                                    for enum in evs.enumeratedValue:
                                        evs_string += str("\n    " + " - " + hex(enum.value) +" : " + enum.name + " : " + enum.description)
                                funcSheet.cell(row=fieldVStart, column=7).value = "{}\n{}".format(field.description, evs_string)
                            else:
                                funcSheet.cell(row=fieldVStart, column=7).value = field.description
                            funcSheet.cell(row=fieldVStart, column=7).alignment = Alignment(wrapText=True)
                            fieldVStart += 1

                        self.regCells.append(registerVStart)
                        endRegister = fieldVStart - 1
                        regPosition = "B{}:G{}".format(registerVStart, endRegister)
                        self.setBorder(funcSheet, regPosition)
                        registerVStart = fieldVStart + 2
                
        for row in self.regCells:
            descriptionRow = row + 2
            funcSheet.row_dimensions[descriptionRow].height = 15


    def createMappingSheet(self):
        pathName = [self.path, self.projectName]
        mappingSheet = self.book.create_sheet("Mapping", 0)
        #mappingSheet = openedBook.active
        
        # Header 
        mappingSheet.cell(row=3, column=1).value = "Function"
        mappingSheet.cell(row=3, column=1).border = self.mediumBorder
        mappingSheet.cell(row=3, column=2).value = "Base Address"
        mappingSheet.cell(row=3, column=2).border = self.mediumBorder
        mappingSheet.cell(row=3, column=3).value = "Local Address"
        mappingSheet.cell(row=3, column=3).border = self.mediumBorder
        mappingSheet.cell(row=3, column=4).value = "Name"
        mappingSheet.cell(row=3, column=4).border = self.mediumBorder
        mappingSheet.cell(row=3, column=5).value = "Description"
        mappingSheet.cell(row=3, column=5).border = self.mediumBorder

        # Freezed Panel
        freezedRow = mappingSheet["E4"]
        mappingSheet.freeze_panes = freezedRow

        # Dimension Group
        mappingSheet.column_dimensions.group('E', 'E', hidden=False)
        mappingSheet.column_dimensions['A'].width = 20
        mappingSheet.column_dimensions['B'].width = 15
        mappingSheet.column_dimensions['C'].width = 15
        mappingSheet.column_dimensions['D'].width = 28
        mappingSheet.column_dimensions['E'].width = 40

        # Populate Mapping
        bits = []
        for memoryMap in self.memoryMaps:
            for addressSpace in memoryMap.addressBlock:
                bits.append(addressSpace.width)

        # OLD ---- bits = [int(addressSpace.width) for addressSpace in data]

        bitStartCol = 6 
        funcRow = 5
        regRow = 5
        line = 0
        rowStartDuplicateReg = 0
        rowEndDuplicateReg = 0
        dimOffset = 1

        # Header
        for bit in reversed(range(max(bits))):
            mappingSheet.cell(row=3, column=bitStartCol).value = bit
            mappingSheet.cell(row=3, column=bitStartCol).border = self.mediumBorder
            fillRegRow = 5 

            for memoryMap in self.memoryMaps:
                for addressSpace in memoryMap.addressBlock:
                    for register in addressSpace.register:
                        for field in register.field:
                            if bit < int(addressSpace.width):
                                mappingSheet.cell(row=fillRegRow, column=bitStartCol).fill = PatternFill("solid", fgColor="EDEDED")
                        fillRegRow += 1
                    fillRegRow += 1 + len(addressSpace.memoryBlockData)
                bitStartCol += 1

        for memoryMap in self.memoryMaps:
            for addressSpace in memoryMap.addressBlock:
                funcHeaderSize =  (len(addressSpace.register) + len(addressSpace.memoryBlockData) + funcRow) if len(addressSpace.register) > 0 else funcRow + 1
                funcBaseAddressSize = len(addressSpace.register) + funcRow if len(addressSpace.register) > 0 else funcRow + 1
                # Function Merged Column
                mappingSheet.merge_cells(start_row=funcRow, start_column=1, end_row = funcHeaderSize - 1, end_column=1) 
                mappingSheet.cell(row=funcRow, column=1).value = addressSpace.name
                mappingSheet.cell(row=funcRow, column=1).border = self.thinBorder
                
                # Base Address Merged Column
                mappingSheet.merge_cells(start_row=funcRow, start_column=2, end_row = funcBaseAddressSize - 1, end_column=2) 
                mappingSheet.cell(row=funcRow, column=2).value = addressSpace.baseAddress
                mappingSheet.cell(row=funcRow, column=2).border = self.thinBorder

                if len(addressSpace.register) > 0:
                    for register in sorted(addressSpace.register, key=lambda register: register.addressOffset):
                        link = "#'{}'!B{}".format(addressSpace.name, self.regCells[line])
                        line += 1
                        if register.dim <= 1:
                            mappingSheet.cell(row=regRow, column=3).value = hex(register.addressOffset)
                            if (regRow >= rowEndDuplicateReg):
                                if (regRow == rowEndDuplicateReg):
                                    mappingSheet.row_dimensions.group(start=rowStartDuplicateReg, end=rowEndDuplicateReg, hidden=True)
                                rowStartDuplicateReg = 0
                                rowEndDuplicateReg = 0
                                dimOffset = 1
                        else:
                            #if previous reg is duplicated finish the collapse
                            if (regRow == rowEndDuplicateReg):
                                mappingSheet.row_dimensions.group(start=rowStartDuplicateReg, end=rowEndDuplicateReg, hidden=True)
                                dimOffset = 1
                            #first handle duplicated regs with offset
                            if (rowStartDuplicateReg == regRow):
                                dimOffset += 1
                            mappingSheet.cell(row=regRow, column=3).value = "({0} + n*{1})\n0 <= n <= {2}".format(register.addressOffset, register.dimOffset, register.dim - 1)
                            rowStartDuplicateReg = regRow + 1
                            rowEndDuplicateReg = regRow + (register.dim - 1)*  dimOffset

                        mappingSheet.cell(row=regRow, column=3).alignment = Alignment(wrapText=True)
                        mappingSheet.cell(row=regRow, column=3).border = self.thinBorder
                        mappingSheet.cell(row=regRow, column=4).value ='=HYPERLINK("{}", "{}")'.format(link, register.name) 
                        mappingSheet.cell(row=regRow, column=4).style = "Hyperlink"
                        mappingSheet.cell(row=regRow, column=4).border = self.thinBorder
                        mappingSheet.cell(row=regRow, column=5).value = register.description
                        mappingSheet.cell(row=regRow, column=5).alignment = Alignment(wrapText=True)
                        mappingSheet.cell(row=regRow, column=5).border = self.thinBorder

                        for field in reversed(register.field):
                            fieldStart = 6
                            posl = field.bitOffset
                            posh = posl + field.bitWidth - 1
                            size = posh - posl + 1
                            if max(bits) - int(addressSpace.width) != 0:
                                mappingSheet.cell(row=regRow, column= (max(bits) - int(addressSpace.width)) + (fieldStart)).border = Border(left=Side(style='medium'))
                                
                            if size == 1:
                                posStart = (max(bits) - int(addressSpace.width)) + (fieldStart + int(addressSpace.width)) - int(posh) - 1
                                mappingSheet.cell(row=regRow, column=posStart).value = field.name
                                mappingSheet.cell(row=regRow, column=posStart).border = self.thinBorder
                                mappingSheet.cell(row=regRow, column=posStart).fill = PatternFill("solid", fgColor="FFFFFF")
                                mappingSheet.cell(row=regRow, column=posStart).alignment = Alignment(wrapText=True)
                            elif size > 1:
                                pos = fieldStart 
                                posStart = ((max(bits) - int(addressSpace.width)) + (fieldStart + int(addressSpace.width))) - int(posh) - 1
                                posEnd = ((max(bits) - int(addressSpace.width)) + (fieldStart + int(addressSpace.width))) - int(posl) - 1
                                mappingSheet.merge_cells(start_row = regRow, 
                                                        start_column = posStart, 
                                                        end_row = regRow, 
                                                        end_column = posEnd)
                                mappingSheet.cell(row=regRow, column=posStart).value = field.name
                                mappingSheet.cell(row=regRow, column=posStart).fill = PatternFill("solid", fgColor="FFFFFF")
                                mappingSheet.cell(row=regRow, column=posStart).alignment = Alignment(wrapText=True)
                            fieldStart += size
                        regRow += 1


                for block in sorted(addressSpace.memoryBlockData, key=lambda block: block.vendorExtensions.baseAddress):
                    mappingSheet.cell(row=regRow, column=2).value = block.vendorExtensions.baseAddress
                    mappingSheet.cell(row=regRow, column=2).border = self.thinBorder
                    localAddress = hex(int(block.vendorExtensions.size)-1)
                    localAddressFormated = "0x{}".format(localAddress[2:].upper())
                    mappingSheet.cell(row=regRow, column=3).value = "0x0\n{}".format(localAddressFormated)
                    mappingSheet.cell(row=regRow, column=3).alignment = Alignment(wrapText=True)
                    mappingSheet.cell(row=regRow, column=3).border = self.thinBorder
                    mappingSheet.cell(row=regRow, column=4).value = block.name
                    mappingSheet.cell(row=regRow, column=4).border = self.thinBorder
                    mappingSheet.cell(row=regRow, column=5).value = block.description
                    mappingSheet.cell(row=regRow, column=5).alignment = Alignment(wrapText=True)
                    mappingSheet.cell(row=regRow, column=5).border = self.thinBorder
                    regRow += 1

                funcRow += len(addressSpace.memoryBlockData) + len(addressSpace.register) + 1
                regRow = funcRow
        
        # Style : align center and merged cell borders

        for row in mappingSheet.rows:
            for i, cell in enumerate(row):
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj

        for merged_cells in mappingSheet.merged_cells.ranges:
            for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                for row in range(merged_cells.min_row, merged_cells.max_row + 1): 
                    mappingSheet.cell(row, col).border = self.thinBorder




path = sys.argv[1]
data = sys.argv[2]

exportExcel = ExportExcel(path, data)
exportExcel.write()
